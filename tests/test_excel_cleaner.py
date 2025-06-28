import json

from openpyxl import Workbook

import excel_cleaner_and_formatter as ecf


def test_excel_cleaner(tmp_path, monkeypatch):
    samples = tmp_path / "samples"
    output = tmp_path / "output"
    samples.mkdir()
    (tmp_path / "output").mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "1"
    ws["B1"] = "O2"
    ws["A2"] = None
    ws["B2"] = "I3"
    wb.save(samples / "test.xlsx")

    monkeypatch.chdir(tmp_path)
    ecf.main()

    json_path = output / "cleaned_data.json"
    assert json_path.exists()
    data = json.loads(json_path.read_text(encoding="utf-8"))
    key = "test.xlsx::Sheet1"
    assert key in data
    assert data[key] == [[1, 2], [-1, -1]]

    csv_files = list(output.glob("test__Sheet1_2x2.csv"))
    assert csv_files and csv_files[0].read_text(encoding="utf-8").startswith(",1,2")


def test_excel_cleaner_dedup_and_range(tmp_path, monkeypatch):
    samples = tmp_path / "samples"
    output = tmp_path / "output"
    samples.mkdir()
    (tmp_path / "output").mkdir()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "1"
    ws["B1"] = "1"  # duplicate
    ws["C1"] = "10"  # out of range for 2x3 grid
    ws["A2"] = "5"
    ws["B2"] = "2"
    ws["C2"] = "3"
    wb.save(samples / "test.xlsx")

    monkeypatch.chdir(tmp_path)
    ecf.main()

    data = json.loads((output / "cleaned_data.json").read_text(encoding="utf-8"))
    key = "test.xlsx::Sheet1"
    assert data[key] == [[1, -1, -1], [5, 2, 3]]
