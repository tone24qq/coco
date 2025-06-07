# app.py
import os, json, logging, zipfile, numpy as np
from io import BytesIO
from fastapi import FastAPI, UploadFile, File, Form, Request
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
import pandas as pd
import asyncio
from analyzer import analyze_board

# 設置日誌
logging.basicConfig(
    format="%(asctime)s %(levelname)-7s [%(name)s] %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# FastAPI App
app = FastAPI()

# 啟動時預處理
@app.on_event("startup")
async def startup_process_samples():
    data_dir = "samples/data"
    output_dir = "samples/output"
    os.makedirs(output_dir, exist_ok=True)

    default_weights = {
        "compute_dynamic_hot_cold_vectorized": 0.2,
        "compute_block_heatmap_vectorized": 0.15,
        "idw_vectorized": 0.15,
        "compute_global_diff_heatmap": 0.1,
        "compute_focus_score": 0.15,
        "detect_skip_patterns": 0.1,
        "compute_difference_trend": 0.05,
        "detect_mirror_sequences": 0.05,
        "connectivity_heatmap": 0.05,
        "sequence_tail_analyzer": 0.05
    }
    return_predictions = False

    for fname in os.listdir(data_dir):
        path = os.path.join(data_dir, fname)
        try:
            # Excel 檔案
            if fname.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                wb = load_workbook(path, data_only=True)
                ws = wb.active
                grid = [[int(cell) if cell is not None and str(cell).isdigit() else -1 for cell in row] for row in ws.iter_rows(values_only=True)]
                grid = np.array(grid)
            # 壓縮包內的 Excel
            elif fname.lower().endswith(".zip"):
                with zipfile.ZipFile(path) as zf:
                    for zi in zf.namelist():
                        if zi.lower().endswith((".xlsx", ".xlsm")):
                            data = zf.read(zi)
                            wb = load_workbook(BytesIO(data), data_only=True)
                            ws = wb.active
                            grid = [[int(cell) if cell is not None and str(cell).isdigit() else -1 for cell in row] for row in ws.iter_rows(values_only=True)]
                            grid = np.array(grid)
            else:
                continue

            # 計算熱力圖
            heatmap_scores, _, _ = await asyncio.to_thread(analyze_board, grid, default_weights, return_predictions)
            out_name = os.path.splitext(fname)[0] + "_heatmap.json"
            with open(os.path.join(output_dir, out_name), "w", encoding="utf-8") as f:
                json.dump({"heatmap": heatmap_scores.tolist()}, f, ensure_ascii=False)
            logger.info(f"Processed {fname}")

        except Exception as e:
            logger.error(f"讀取 {fname} 失敗：{e}")

def process_grid(grid: np.ndarray, weights: dict, return_predictions: bool, target_num: int = None, json_heatmap: str = None) -> dict:
    if grid.size == 0 or grid.shape[0] < 4 or grid.shape[1] < 5 or grid.shape[0] > 20 or grid.shape[1] > 20:
        return {"error": "網格為空或超出 4x5 至 20x20 限制"}
    final_score, final_pred, best_pos = analyze_board(grid, weights, return_predictions, target_num, json_heatmap)
    result = {"heatmap": final_score.tolist()}
    if return_predictions and final_pred is not None:
        result["prediction"] = final_pred.tolist()
    if best_pos:
        result["best_position"] = best_pos[0] if best_pos else None
        result["top_3"] = best_pos[1:] if len(best_pos) > 1 else []
    return result

def parse_weights(weights: str) -> dict:
    if weights:
        try:
            return json.loads(weights)
        except json.JSONDecodeError:
            raise ValueError("無効的 JSON 格式")
    return {
        "compute_dynamic_hot_cold_vectorized": 0.2,
        "compute_block_heatmap_vectorized": 0.15,
        "idw_vectorized": 0.15,
        "compute_global_diff_heatmap": 0.1,
        "compute_focus_score": 0.15,
        "detect_skip_patterns": 0.1,
        "compute_difference_trend": 0.05,
        "detect_mirror_sequences": 0.05,
        "connectivity_heatmap": 0.05,
        "sequence_tail_analyzer": 0.05
    }

def load_file_content(filepath: str) -> list:
    ext = os.path.splitext(filepath)[1].lower()
    grids = []
    if ext == ".json":
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        grids = [np.array(data)]
    elif ext == ".csv":
        df = pd.read_csv(filepath, header=None, dtype=str, keep_default_na=False)
        cleaned = [[int(c.replace('O','0').replace('I','1')) if c.isdigit() else -1 for c in row] for row in df.values]
        grids.append(np.array(cleaned))
    else:  # .xls, .xlsx
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows():
                cleaned_row = []
                for cell in row:
                    val = str(cell.value) if cell.value is not None else ""
                    val = val.replace('O', '0').replace('I', '1')
                    cleaned_row.append(int(val) if val.isdigit() else -1)
                rows.append(cleaned_row)
            grids.append(np.array(rows))
    return grids

@app.post("/analyze/")
async def analyze(file: UploadFile = File(...), weights: str = Form(None), mode: str = Form("predict"), target_num: int = Form(None), json_heatmap: str = Form(None)):
    filename = file.filename.lower()
    if not filename.endswith((".xls", ".xlsx", ".json", ".csv")):
        return JSONResponse(status_code=380, content={"error": "不支援的檔案格式"})
    content = await file.read()
    grids = []
    if filename.endswith(".json"):
        try:
            data = json.loads(content.decode("utf-8"))
            grids = [np.array(data)]
        except json.JSONDecodeError:
            return JSONResponse(status_code=380, content={"error": "無効的 JSON 格式"})
    elif filename.endswith(".csv"):
        df = pd.read_csv(BytesIO(content), header=None, dtype=str, keep_default_na=False)
        cleaned = [[int(c.replace('O','0').replace('I','1')) if c.isdigit() else -1 for c in row] for row in df.values]
        grids.append(np.array(cleaned))
    else:  # .xls, .xlsx
        try:
            wb = load_workbook(BytesIO(content), data_only=True)
            grids = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows():
                    cleaned_row = []
                    for cell in row:
                        val = str(cell.value) if cell.value is not None else ""
                        val = val.replace('O', '0').replace('I', '1')
                        cleaned_row.append(int(val) if val.isdigit() else -1)
                    rows.append(cleaned_row)
                grids.append(np.array(rows))
        except Exception as e:
            return JSONResponse(status_code=380, content={"error": f"Excel 讀取失敗: {str(e)}"})
    try:
        w_dict = parse_weights(weights)
    except ValueError as e:
        return JSONResponse(status_code=380, content={"error": str(e)})
    return_predictions = (mode == "predict")
    results = []
    for idx, grid in enumerate(grids):
        result = await asyncio.to_thread(process_grid, grid, w_dict, return_predictions, target_num, json_heatmap)
        result["sheet"] = idx + 1
        results.append(result)
    return JSONResponse(content={"results": results})

@app.post("/analyze-batch/")
async def analyze_batch(file: UploadFile = File(...), weights: str = Form(None), mode: str = Form("predict"), target_num: int = Form(None), json_heatmap: str = Form(None)):
    filename = file.filename.lower()
    if not filename.endswith(".zip"):
        return JSONResponse(status_code=380, content={"error": "請上傳 ZIP 檔案"})
    content = await file.read()
    results = []
    try:
        w_dict = parse_weights(weights)
    except ValueError as e:
        return JSONResponse(status_code=380, content={"error": str(e)})
    return_predictions = (mode == "predict")
    try:
        with zipfile.ZipFile(BytesIO(content)) as z:
            for zip_info in z.infolist():
                name = zip_info.filename
                if name.endswith((".xls", ".xlsx")):
                    with z.open(zip_info) as f:
                        wb = load_workbook(BytesIO(f.read()), data_only=True)
                        file_results = {"filename": name, "sheets": []}
                        for sheet in wb.sheetnames:
                            ws = wb[sheet]
                            rows = []
                            for row in ws.iter_rows():
                                cleaned_row = []
                                for cell in row:
                                    val = str(cell.value) if cell.value is not None else ""
                                    val = val.replace('O', '0').replace('I', '1')
                                    cleaned_row.append(int(val) if val.isdigit() else -1)
                                rows.append(cleaned_row)
                            grid = np.array(rows)
                            result = await asyncio.to_thread(process_grid, grid, w_dict, return_predictions, target_num, json_heatmap)
                            result["sheet"] = sheet
                            file_results["sheets"].append(result)
                        results.append(file_results)
                elif name.endswith(".json"):
                    with z.open(zip_info) as f:
                        try:
                            data = json.load(f)
                            grid = np.array(data)
                        except json.JSONDecodeError:
                            results.append({"filename": name, "error": "無効的 JSON 格式"})
                            continue
                        file_results = {"filename": name, "sheets": []}
                        result = await asyncio.to_thread(process_grid, grid, w_dict, return_predictions, target_num, json_heatmap)
                        result["sheet"] = "data"
                        file_results["sheets"].append(result)
                        results.append(file_results)
                elif name.endswith(".csv"):
                    with z.open(zip_info) as f:
                        df = pd.read_csv(BytesIO(f.read()), header=None, dtype=str, keep_default_na=False)
                        cleaned = [[int(c.replace('O','0').replace('I','1')) if c.isdigit() else -1 for c in row] for row in df.values]
                        grid = np.array(cleaned)
                        file_results = {"filename": name, "sheets": []}
                        result = await asyncio.to_thread(process_grid, grid, w_dict, return_predictions, target_num, json_heatmap)
                        result["sheet"] = "data"
                        file_results["sheets"].append(result)
                        results.append(file_results)
    except zipfile.BadZipFile:
        return JSONResponse(status_code=380, content={"error": "無効的 ZIP 檔案"})
    return JSONResponse(content={"results": results})

@app.post("/analyze-folder/")
async def analyze_folder(weights: str = Form(None), mode: str = Form("predict"), target_num: int = Form(None), json_heatmap: str = Form(None)):
    folder_path = "samples/data/"
    if not os.path.exists(folder_path):
        return JSONResponse(content={"error": f"資料夾 {folder_path} 不存在"})
    try:
        w_dict = parse_weights(weights)
    except Exception as e:
        return JSONResponse(status_code=380, content={"error": str(e)})
    return_predictions = (mode == "predict")
    results = []
    for idx, filename in enumerate(os.listdir(folder_path)):
        filepath = os.path.join(folder_path, filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".json", ".csv", ".xls", ".xlsx"]:
            continue
        try:
            grids = load_file_content(filepath)
            file_results = {"filename": filename, "sheets": []}
            for sidx, grid in enumerate(grids):
                result = await asyncio.to_thread(process_grid, grid, w_dict, return_predictions, target_num, json_heatmap)
                result["sheet"] = f"Sheet{sidx+1}" if ext in [".xls", ".xlsx"] else "data"
                file_results["sheets"].append(result)
            results.append(file_results)
        except Exception as e:
            logger.error(f"Error processing {filename}: {str(e)}")
            results.append({"filename": filename, "error": f"檔案處理失敗: {str(e)}"})
    return JSONResponse(content={"results": results})

@app.get("/")
async def root():
    return JSONResponse(status_code=200, content={"status": "running"})

@app.api_route("/{full_path:path}", methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS", "HEAD"])
async def catch_all(request: Request, full_path: str):
    logger.debug(f"Catch-all for path: {request.method} {full_path}")
    return JSONResponse(status_code=200, content={"status": "running"})