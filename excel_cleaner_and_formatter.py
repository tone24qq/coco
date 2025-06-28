import csv
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def clean_cell(value: Any) -> Any:
    """Clean individual cell value according to rules."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = s.replace("O", "0").replace("I", "1")
    s_up = s.upper()
    if s_up.isdigit():
        return int(s_up)
    return ""


def process_sheet(ws) -> List[List[Any]]:
    data: List[List[Any]] = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for r in range(1, max_row + 1):
        row: List[Any] = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            row.append(clean_cell(val))
        data.append(row)
    return data


def save_visual_csv(data: List[List[Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not data:
        with open(path, "w", newline="", encoding="utf-8") as f:
            f.write("")
        return
    header = [""] + [str(i + 1) for i in range(len(data[0]))]
    rows = []
    for idx, row in enumerate(data, start=1):
        rows.append([str(idx)] + [str(v) if v != "" else "" for v in row])
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def to_int_grid(data: List[List[Any]]) -> List[List[int]]:
    """Convert cleaned cell values to an integer grid.

    Only values in the range ``1..N`` (where ``N`` is ``rows*cols``) are kept
    and must not repeat. All other cells are set to ``-1``. This prevents
    stray dates or duplicated numbers from polluting the dataset.
    """

    rows = len(data)
    cols = len(data[0]) if rows > 0 else 0
    max_val = rows * cols
    seen: set[int] = set()
    result: List[List[int]] = []

    for row in data:
        row_int: List[int] = []
        for val in row:
            if isinstance(val, int):
                num = val
            elif isinstance(val, str) and val.isdigit():
                num = int(val)
            else:
                num = -1

            if 1 <= num <= max_val and num not in seen:
                row_int.append(num)
                seen.add(num)
            else:
                row_int.append(-1)

        result.append(row_int)

    return result


def main() -> None:
    samples = Path("samples")
    output = Path("output")
    output.mkdir(exist_ok=True)
    result: Dict[str, List[List[int]]] = {}

    for xlsx in samples.glob("*.xlsx"):
        wb = load_workbook(xlsx, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            data = process_sheet(ws)
            csv_name = f"{xlsx.stem}__{name}_{len(data)}x{len(data[0])}.csv"
            save_visual_csv(data, output / csv_name)
            result[f"{xlsx.name}::{name}"] = to_int_grid(data)
    with open(output / "cleaned_data.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)


if __name__ == "__main__":
    main()
