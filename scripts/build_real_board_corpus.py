from __future__ import annotations

import argparse
import sys
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.safe_io import SafeWriteConfig, write_jsonl_records_safe  # noqa: E402


def validate_grid(rows: int, cols: int, grid: List[List[Optional[int]]]) -> Tuple[bool, str]:
    if rows <= 0 or cols <= 0:
        return False, "invalid_shape"
    flat = [v for row in grid for v in row if v is not None]
    if len(flat) != rows * cols:
        return False, "partial_board"
    expected = set(range(1, rows * cols + 1))
    got = set(flat)
    if got != expected:
        return False, "not_permutation_1_to_n"
    return True, "ok"


def _to_int_or_none(value: Any) -> Optional[int]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, float) and float(value).is_integer():
        return int(value)
    return None


def _record_from_grid(
    *,
    board_id: str,
    source_file: str,
    issue_id: str,
    grid: List[List[Optional[int]]],
    source: str,
) -> Dict[str, Any]:
    rows_n, cols_n = len(grid), (len(grid[0]) if grid else 0)
    ok, reason = validate_grid(rows_n, cols_n, grid)
    return {
        "board_id": board_id,
        "lineage_id": board_id,
        "rows": rows_n,
        "cols": cols_n,
        "size_class": f"{rows_n}x{cols_n}",
        "grid": grid,
        "is_full_board": bool(ok),
        "validation_reason": reason,
        "source_type": "real",
        "is_real": True,
        "source": source,
        "source_file": source_file,
        "issue_id": issue_id,
        "group_id": board_id,
    }


def scan_xlsx(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    records: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)
        if max_row <= 0 or max_col <= 0:
            continue
        grid: List[List[Optional[int]]] = []
        for r in range(1, max_row + 1):
            row_vals: List[Optional[int]] = []
            for c in range(1, max_col + 1):
                row_vals.append(_to_int_or_none(ws.cell(r, c).value))
            grid.append(row_vals)

        while grid and all(v is None for v in grid[-1]):
            grid.pop()
        if not grid:
            continue
        max_used_col = max(
            (idx for row in grid for idx, v in enumerate(row, start=1) if v is not None),
            default=0,
        )
        if max_used_col == 0:
            continue
        grid = [row[:max_used_col] for row in grid]
        board_id = f"{path.stem}:{ws.title}"
        records.append(
            _record_from_grid(
                board_id=board_id,
                source_file=path.name,
                issue_id=path.stem,
                grid=grid,
                source="xlsx_scan",
            )
        )
    return records


def scan_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as fh:
        for i, line in enumerate(fh):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            grid = obj.get("grid")
            if not isinstance(grid, list) or not grid:
                continue
            normalized: List[List[Optional[int]]] = [[_to_int_or_none(v) for v in row] for row in grid]
            board_id = str(obj.get("board_id") or f"{path.stem}:{i}")
            rows.append(
                _record_from_grid(
                    board_id=board_id,
                    source_file=path.name,
                    issue_id=path.stem,
                    grid=normalized,
                    source="jsonl_scan",
                )
            )
    return rows


def scan_parquet(path: Path) -> List[Dict[str, Any]]:
    df = pd.read_parquet(path)
    rows: List[Dict[str, Any]] = []
    for i, obj in enumerate(df.to_dict(orient="records")):
        grid = obj.get("grid")
        if isinstance(grid, np.ndarray):
            grid = grid.tolist()
        if isinstance(grid, str):
            try:
                grid = json.loads(grid)
            except Exception:
                continue
        if not isinstance(grid, list) or not grid:
            continue
        normalized: List[List[Optional[int]]] = [[_to_int_or_none(v) for v in row] for row in grid]
        board_id = str(obj.get("board_id") or f"{path.stem}:{i}")
        rows.append(
            _record_from_grid(
                board_id=board_id,
                source_file=path.name,
                issue_id=path.stem,
                grid=normalized,
                source="parquet_scan",
            )
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default=".")
    parser.add_argument("--glob", default="*")
    parser.add_argument("--output", default="data/full_boards/full_board_corpus.jsonl")
    parser.add_argument("--partial-meta", default="data/full_boards/partial_real_board_metadata.jsonl")
    parser.add_argument("--audit", default="reports/full_board_corpus_audit.json")
    parser.add_argument("--preview-dir", default="reports/root_xlsx_previews")
    parser.add_argument("--max-file-mb", type=int, default=100)
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    files = sorted(input_dir.rglob(args.glob))

    all_records: List[Dict[str, Any]] = []
    for file in files:
        suffix = file.suffix.lower()
        if suffix == ".xlsx":
            all_records.extend(scan_xlsx(file))
        elif suffix == ".jsonl":
            all_records.extend(scan_jsonl(file))
        elif suffix == ".parquet":
            all_records.extend(scan_parquet(file))

    full = [r for r in all_records if r["is_full_board"]]
    partial = [r for r in all_records if not r["is_full_board"]]

    write_jsonl_records_safe(
        full,
        Path(args.output),
        config=SafeWriteConfig(
            max_file_mb=args.max_file_mb,
            producer_script="scripts/build_real_board_corpus.py",
        ),
    )
    write_jsonl_records_safe(
        partial,
        Path(args.partial_meta),
        config=SafeWriteConfig(
            max_file_mb=args.max_file_mb,
            producer_script="scripts/build_real_board_corpus.py",
        ),
    )

    preview_dir = Path(args.preview_dir)
    preview_dir.mkdir(parents=True, exist_ok=True)
    for rec in full[:50]:
        p = preview_dir / f"{rec['board_id'].replace(':', '__')}.txt"
        grid = rec["grid"]
        p.write_text("\n".join(["\t".join(str(v) for v in row) for row in grid]), encoding="utf-8")

    per_size: Dict[str, int] = {}
    for row in full:
        per_size[row["size_class"]] = per_size.get(row["size_class"], 0) + 1

    audit = {
        "status": "ok",
        "input_dir": str(input_dir),
        "glob": args.glob,
        "scanned_files": [str(f) for f in files],
        "full_board_count": len(full),
        "partial_board_count": len(partial),
        "per_size": per_size,
    }
    Path(args.audit).parent.mkdir(parents=True, exist_ok=True)
    Path(args.audit).write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"full": len(full), "partial": len(partial)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
