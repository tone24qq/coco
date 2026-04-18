from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

DIGIT_RE = re.compile(r"^\d+$")
SIZE_HINT_RE = re.compile(r"(?<!\d)(\d{1,2})\s*[xX]\s*(\d{1,2})(?!\d)")
COMMON_SHAPES: Tuple[Tuple[int, int], ...] = ((4, 5), (6, 10), (8, 10), (10, 10), (10, 12), (10, 16))


def read_sheet_as_strings(ws) -> List[List[str]]:
    table: List[List[str]] = []
    for row in ws.iter_rows():
        table.append(["" if cell.value is None else str(cell.value) for cell in row])
    if not table:
        return []
    width = max(len(r) for r in table)
    return [r + [""] * (width - len(r)) for r in table]


def normalize_token(raw: str) -> str:
    s = (raw or "").strip().upper()
    s = s.replace("O", "0").replace("I", "1")
    if s.endswith(".0") and DIGIT_RE.fullmatch(s[:-2]):
        s = s[:-2]
    return s


def parse_int_token(raw: str) -> Optional[int]:
    s = normalize_token(raw)
    if not s or not DIGIT_RE.fullmatch(s):
        return None
    try:
        return int(s)
    except Exception:
        return None


def board_signature(grid: Sequence[Sequence[int]]) -> str:
    return ",".join(str(v) for row in grid for v in row)


def validate_full_permutation(flat: Sequence[int], n: int) -> Tuple[bool, str]:
    if len(flat) != n:
        return False, "length_mismatch"
    missing = sorted(set(range(1, n + 1)) - set(flat))
    duplicates = sorted(v for v, cnt in Counter(flat).items() if cnt > 1)
    out_of_range = sorted(v for v in flat if v < 1 or v > n)
    if out_of_range:
        return False, f"out_of_range_values:{out_of_range[:8]}"
    if duplicates:
        return False, f"duplicate_values:{duplicates[:8]}"
    if missing:
        return False, f"missing_values:{missing[:8]}"
    return True, "ok"


def matrix_to_visual(table: Sequence[Sequence[str]]) -> str:
    if not table:
        return "[empty]"
    cols = max(len(row) for row in table)
    widths = [0] * cols
    for c in range(cols):
        widths[c] = max(len(str(row[c])) if c < len(row) else 0 for row in table)
        widths[c] = max(widths[c], len(f"C{c+1}"), 1)

    def fmt_row(values: Sequence[str], row_name: str) -> str:
        cells = []
        for c in range(cols):
            v = values[c] if c < len(values) else ""
            cells.append(str(v).rjust(widths[c]))
        return row_name.rjust(4) + " | " + " ".join(cells)

    header = "     | " + " ".join(f"C{c+1}".rjust(widths[c]) for c in range(cols))
    sep = "-" * len(header)
    lines = [header, sep]
    for r, row in enumerate(table, start=1):
        lines.append(fmt_row(row, f"R{r}"))
    return "\n".join(lines)


def extract_submatrix(table: Sequence[Sequence[str]], top: int, left: int, rows: int, cols: int) -> List[List[str]]:
    return [list(r[left : left + cols]) for r in table[top : top + rows]]


def try_parse_board(sub: Sequence[Sequence[str]], rows: int, cols: int) -> Tuple[Optional[List[List[int]]], str]:
    if len(sub) != rows:
        return None, "rows_mismatch"
    out: List[List[int]] = []
    non_int_cells: List[str] = []
    for ridx, row in enumerate(sub, start=1):
        if len(row) != cols:
            return None, "cols_mismatch"
        parsed_row: List[int] = []
        for cidx, cell in enumerate(row, start=1):
            value = parse_int_token(cell)
            if value is None:
                non_int_cells.append(f"R{ridx}C{cidx}")
                parsed_row.append(-999999)
                continue
            parsed_row.append(value)
        out.append(parsed_row)

    if non_int_cells:
        return None, f"non_integer_cells:{non_int_cells[:6]}"

    flat = [v for row in out for v in row]
    ok, reason = validate_full_permutation(flat, rows * cols)
    if not ok:
        return None, reason
    return out, "ok"


def _factor_shapes(n: int) -> List[Tuple[int, int]]:
    out: List[Tuple[int, int]] = []
    if n <= 0:
        return out
    for r in range(2, int(n**0.5) + 1):
        if n % r == 0:
            c = n // r
            out.append((r, c))
            out.append((c, r))
    return out


def _shapes_from_filename(name: str) -> List[Tuple[int, int]]:
    found: List[Tuple[int, int]] = []
    for m in SIZE_HINT_RE.finditer(name):
        shape = (int(m.group(1)), int(m.group(2)))
        if shape not in found:
            found.append(shape)
    return found


def _infer_shapes(table: Sequence[Sequence[str]], filename: str) -> List[Tuple[int, int]]:
    sheet_rows = len(table)
    sheet_cols = max((len(r) for r in table), default=0)

    int_values: List[int] = []
    for row in table:
        for cell in row:
            v = parse_int_token(cell)
            if v is not None and v > 0:
                int_values.append(v)

    inferred: List[Tuple[int, int]] = []
    for n in {max(int_values) if int_values else 0, len(int_values)}:
        for shape in _factor_shapes(n):
            if shape not in inferred:
                inferred.append(shape)

    merged: List[Tuple[int, int]] = []
    for shape in _shapes_from_filename(filename) + list(COMMON_SHAPES) + inferred:
        r, c = shape
        if r < 2 or c < 2:
            continue
        if r > sheet_rows or c > sheet_cols:
            continue
        if shape not in merged:
            merged.append(shape)
    return merged


def scan_sheet_for_boards(
    table: Sequence[Sequence[str]],
    shapes: Sequence[Tuple[int, int]],
) -> Tuple[List[Tuple[int, int, int, int, List[List[int]]]], Counter[str]]:
    if not table:
        return [], Counter({"empty_sheet": 1})

    total_rows = len(table)
    total_cols = max(len(r) for r in table)
    seen_grids = set()
    found: List[Tuple[int, int, int, int, List[List[int]]]] = []
    rejects: Counter[str] = Counter()

    for rows, cols in shapes:
        if total_rows < rows or total_cols < cols:
            continue
        for top in range(total_rows - rows + 1):
            for left in range(total_cols - cols + 1):
                sub = extract_submatrix(table, top, left, rows, cols)
                parsed, reason = try_parse_board(sub, rows, cols)
                if parsed is None:
                    rejects[f"{rows}x{cols}:{reason}"] += 1
                    continue
                key = (rows, cols, board_signature(parsed))
                if key in seen_grids:
                    rejects[f"{rows}x{cols}:duplicate_within_sheet"] += 1
                    continue
                seen_grids.add(key)
                found.append((top, left, rows, cols, parsed))
    return found, rejects


def preview_filename(xlsx_name: str, sheet_name: str, top: int, left: int, rows: int, cols: int) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", f"{Path(xlsx_name).stem}__{sheet_name}__r{top+1}_c{left+1}__{rows}x{cols}")
    return safe + ".txt"


def build_corpus(input_dir: Path, preview_dir: Optional[Path]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    files = sorted([p for p in input_dir.glob("*.xlsx") if not p.name.startswith("~$")])
    records: List[Dict[str, Any]] = []
    size_counts: Counter[str] = Counter()
    audit: Dict[str, Any] = {
        "status": "ok",
        "files_checked": [str(p) for p in files],
        "accepted": [],
        "rejected": [],
        "size_counts": {},
        "errors": [],
        "files": [],
    }
    seen_board_keys = set()

    if preview_dir is not None:
        preview_dir.mkdir(parents=True, exist_ok=True)

    for xlsx_path in files:
        try:
            wb = load_workbook(xlsx_path, data_only=True)
        except Exception as exc:
            audit["errors"].append({"file": xlsx_path.name, "error": f"open_failed: {exc}"})
            continue

        file_hits = 0
        file_report: Dict[str, Any] = {"file": xlsx_path.name, "sheets": [], "accepted_total": 0, "rejected_total": 0}

        for sheet_index, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            try:
                table = read_sheet_as_strings(ws)
            except Exception as exc:
                audit["errors"].append({"file": xlsx_path.name, "sheet": sheet_name, "error": f"read_failed: {exc}"})
                continue

            inferred_shapes = _infer_shapes(table, xlsx_path.name)
            found_boards, reject_reasons = scan_sheet_for_boards(table, inferred_shapes)
            sheet_hits = 0
            sheet_rejects = int(sum(reject_reasons.values()))

            for top, left, rows, cols, grid in found_boards:
                board_key = (rows, cols, board_signature(grid))
                if board_key in seen_board_keys:
                    sheet_rejects += 1
                    continue
                seen_board_keys.add(board_key)

                board_id = f"xlsx:{xlsx_path.stem}:{sheet_name}:{top+1}:{left+1}:{rows}x{cols}"
                record = {
                    "board_id": board_id,
                    "rows": rows,
                    "cols": cols,
                    "size_class": f"{rows}x{cols}",
                    "board_size": rows * cols,
                    "grid": grid,
                    "source": "root_xlsx_import",
                    "source_file": xlsx_path.name,
                    "issue_id": f"{xlsx_path.stem}:{sheet_name}",
                    "group_id": board_id,
                    "is_real": True,
                    "source_type": "real",
                    "order_index": len(records),
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "top": top + 1,
                    "left": left + 1,
                }
                records.append(record)
                size_counts[record["size_class"]] += 1
                sheet_hits += 1
                file_hits += 1

                preview_path = None
                if preview_dir is not None:
                    preview_path = preview_dir / preview_filename(xlsx_path.name, sheet_name, top, left, rows, cols)
                    preview_text = matrix_to_visual([[str(v) for v in row] for row in grid])
                    preview_path.write_text(preview_text, encoding="utf-8")

                audit["accepted"].append(
                    {
                        "file": xlsx_path.name,
                        "sheet": sheet_name,
                        "top": top + 1,
                        "left": left + 1,
                        "rows": rows,
                        "cols": cols,
                        "board_id": board_id,
                        "preview": str(preview_path) if preview_path else None,
                    }
                )

            if sheet_hits == 0:
                audit["rejected"].append(
                    {
                        "file": xlsx_path.name,
                        "sheet": sheet_name,
                        "reason": "no_valid_full_permutation_found_for_inferred_shapes",
                        "inferred_shapes": [f"{r}x{c}" for r, c in inferred_shapes],
                        "top_rejected_reasons": reject_reasons.most_common(10),
                    }
                )

            file_report["sheets"].append(
                {
                    "sheet_name": sheet_name,
                    "inferred_shapes": [f"{r}x{c}" for r, c in inferred_shapes],
                    "accepted_boards": sheet_hits,
                    "rejected_windows": sheet_rejects,
                    "top_rejected_reasons": reject_reasons.most_common(10),
                }
            )
            file_report["accepted_total"] += sheet_hits
            file_report["rejected_total"] += sheet_rejects

        if file_hits == 0:
            audit["errors"].append({"file": xlsx_path.name, "error": "no_valid_permutation_board_found_in_file"})

        audit["files"].append(file_report)

    audit["size_counts"] = dict(size_counts)
    audit["board_count"] = len(records)
    return records, audit


def main() -> None:
    parser = argparse.ArgumentParser(description="Build multi-size board corpus from root xlsx files.")
    parser.add_argument("--input-dir", default=".")
    parser.add_argument("--output", default="data/full_boards/full_board_corpus_80.jsonl")
    parser.add_argument("--audit", default="reports/full_board_corpus_audit.json")
    parser.add_argument("--preview-dir", default="reports/root_xlsx_previews")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    audit_path = Path(args.audit)
    preview_dir = Path(args.preview_dir) if args.preview_dir else None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.parent.mkdir(parents=True, exist_ok=True)

    records, audit = build_corpus(input_dir, preview_dir)

    with output_path.open("w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    audit["output"] = str(output_path)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({"board_count": len(records), "size_counts": audit["size_counts"], "output": str(output_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
